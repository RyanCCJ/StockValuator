"""Import API routes for trades and cash transactions."""

import csv
import io
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from fastapi import APIRouter, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from pydantic import ValidationError

from src.api.deps import CurrentUser, DbSession
from src.models.trade import TradeType
from src.models.cash import CashTransactionType
from src.schemas.trade import TradeCreate
from src.schemas.cash import CashTransactionCreate
from src.services.trade_service import create_trade
from src.services.cash_service import create_cash_transaction

router = APIRouter(prefix="/import", tags=["import"])


class ImportResult:
    """Result of an import operation."""
    
    def __init__(self):
        self.success_count = 0
        self.errors: list[dict[str, Any]] = []


@router.post("/trades")
async def import_trades(
    current_user: CurrentUser,
    db: DbSession,
    file: UploadFile = File(...),
):
    """Import trades from CSV or XLSX file."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    ext = file.filename.lower().split(".")[-1]
    if ext not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="File must be .csv or .xlsx")
    
    content = await file.read()
    rows = _parse_file(content, ext)
    
    result = ImportResult()
    
    for i, row in enumerate(rows, start=2):  # Start at 2 (row 1 is header)
        try:
            trade_data = _parse_trade_row(row)
            await create_trade(db, current_user.id, trade_data)
            result.success_count += 1
        except Exception as e:
            result.errors.append({"row": i, "error": str(e)})
    
    await db.commit()
    
    return {
        "success_count": result.success_count,
        "error_count": len(result.errors),
        "errors": result.errors[:20],  # Limit errors returned
    }


@router.post("/cash")
async def import_cash(
    current_user: CurrentUser,
    db: DbSession,
    file: UploadFile = File(...),
):
    """Import cash transactions from CSV or XLSX file."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    ext = file.filename.lower().split(".")[-1]
    if ext not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="File must be .csv or .xlsx")
    
    content = await file.read()
    rows = _parse_file(content, ext)
    
    result = ImportResult()
    
    for i, row in enumerate(rows, start=2):
        try:
            cash_data = _parse_cash_row(row)
            await create_cash_transaction(db, current_user.id, cash_data)
            result.success_count += 1
        except Exception as e:
            result.errors.append({"row": i, "error": str(e)})
    
    await db.commit()
    
    return {
        "success_count": result.success_count,
        "error_count": len(result.errors),
        "errors": result.errors[:20],
    }


def _parse_file(content: bytes, ext: str) -> list[dict[str, str]]:
    """Parse CSV or XLSX file content into list of row dicts."""
    if ext == "csv":
        text = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)
    else:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).lower().strip() if h else "" for h in next(rows_iter)]
        result = []
        for row in rows_iter:
            row_dict = {}
            for j, val in enumerate(row):
                if j < len(headers) and headers[j]:
                    row_dict[headers[j]] = str(val) if val is not None else ""
            result.append(row_dict)
        return result


def _parse_trade_row(row: dict[str, str]) -> TradeCreate:
    """Parse a row dict into TradeCreate schema."""
    # Normalize keys to lowercase
    row = {k.lower().strip(): v for k, v in row.items()}
    
    date_str = row.get("date", "")
    try:
        date = datetime.fromisoformat(date_str.replace(" ", "T"))
    except ValueError:
        # Try common formats
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"]:
            try:
                date = datetime.strptime(date_str, fmt)
                break
            except ValueError:
                continue
        else:
            raise ValueError(f"Invalid date format: {date_str}")
    
    type_str = row.get("type", "").lower()
    if type_str not in ("buy", "sell"):
        raise ValueError(f"Invalid trade type: {type_str}")
    
    return TradeCreate(
        symbol=row.get("symbol", "").upper(),
        date=date,
        type=TradeType(type_str),
        price=Decimal(row.get("price", "0")),
        quantity=Decimal(row.get("quantity", "0")),
        fees=Decimal(row.get("fees", "0") or "0"),
        currency=row.get("currency", "USD").upper() or "USD",
        notes=row.get("notes") or None,
    )


def _parse_cash_row(row: dict[str, str]) -> CashTransactionCreate:
    """Parse a row dict into CashTransactionCreate schema."""
    row = {k.lower().strip(): v for k, v in row.items()}
    
    date_str = row.get("date", "")
    try:
        date = datetime.fromisoformat(date_str.replace(" ", "T"))
    except ValueError:
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"]:
            try:
                date = datetime.strptime(date_str, fmt)
                break
            except ValueError:
                continue
        else:
            raise ValueError(f"Invalid date format: {date_str}")
    
    type_str = row.get("type", "").lower()
    if type_str not in ("deposit", "withdrawal"):
        raise ValueError(f"Invalid transaction type: {type_str}")
    
    return CashTransactionCreate(
        date=date,
        type=CashTransactionType(type_str),
        amount=Decimal(row.get("amount", "0")),
        currency=row.get("currency", "USD").upper() or "USD",
        notes=row.get("notes") or None,
    )
